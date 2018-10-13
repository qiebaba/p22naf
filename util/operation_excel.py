# 封装EXCEL文件的操作方法
from openpyxl import load_workbook
import time


class ParseExcel:
    def __init__(self):
        self.workbook = None
        self.excel_file = None
        self.sheet = None

    # 打开excel文件,返回文件对象
    def load_workbook(self, excel_path):
        try:
            self.workbook = load_workbook(excel_path)
        except Exception as e:
            raise e
        self.excel_file = excel_path
        return self.workbook

    # 根据sheet名获取sheet对象
    def get_sheet_by_name(self, sheet_name):
        try:
            self.sheet = self.workbook[sheet_name]
            return self.sheet
        except Exception as e:
            raise e

    # 获取结束行号
    def get_row_number(self):
        try:
            row_num = self.sheet.max_row
            return row_num
        except Exception as e:
            raise e

    # 获取结束列号
    def get_cols_number(self):
        try:
            col_num = self.sheet.max_column
            return col_num
        except Exception as e:
            raise e

    # 获取莫一行的数据
    def get_row_data(self, row_no):
        try:
            row_data = self.sheet[row_no]
            return row_data
        except Exception as e:
            raise e

    # 获取某一列的数据
    def get_column_data(self, col_no):
        try:
            col_data = self.sheet[col_no]
            return col_data
        except Exception as e:
            raise e

    # 根据单元格位置获取单元格的值
    def get_cell_value(self, rowNo, colsNo):
        try:
            self.cell_value = self.sheet.cell(rowNo, colsNo)
            return self.cell_value.value
        except Exception as e:
            raise e

    # 保存excel文件
    def save_excel(self, excel_file):
        try:
            self.workbook.save(excel_file)
        except Exception as e:
            raise e

    # 根据单元格的坐标向单元格写入数据
    def write_cell(self, content, row_no=None, cols_no=None):
        try:
            self.get_cell_value(row_no, cols_no)
            self.cell_value.value = content
            self.save_excel(self.excel_file)
        except Exception as e:
            raise e

    # 写入当前时间
    def write_current_time(self, row_no=None, cols_no=None):
        try:
            now = int(time.time())
            time_array = time.localtime(now)
            current_time = time.strftime("%Y-%m-%d%H:%M:%S", time_array)
            self.cell_value.value = current_time
            self.save_excel(self.excel_file)
        except Exception as e:
            raise e






if __name__ == '__main__':
    pe = ParseExcel()
    pe.load_workbook(r"..\dataconfig\testdata.xlsx")
    sheet = pe.get_sheet_by_name('sheet1')
    print(pe.get_row_number())
    print(pe.get_cols_number())
    print(pe.get_row_data(1))
    print(pe.get_column_data('A'))

    print(pe.get_cell_value(1, 2))

    print(pe.id_get_row_num("xxb-01"))

    print(pe.id_get_row_data("xxb-01"))
